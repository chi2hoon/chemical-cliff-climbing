import pandas as pd
import openpyxl
from datetime import datetime
import os
from pathlib import Path

# --- Configuration ---
# The name of the specific Excel file to process.
# Assumed to be located in the 'sar_raw' subdirectory relative to this file.
EXCEL_FILE_NAME = os.path.join("sar_raw", "1020170094694_extracted_250611.xlsx")

# The directory where all processed CSV files will be saved.
OUTPUT_DIR = "processed_excel_data"

# --- Low-Level Helper Functions ---

def is_row_empty(row_data):
    """
    Checks if a row (represented as a list of values) is effectively empty.
    A row is considered empty if all its cell values are None or empty strings after stripping whitespace.

    Args:
        row_data (list): A list of cell values.

    Returns:
        bool: True if the row is empty, False otherwise.
    """
    # Check if the value is None or an empty string after stripping whitespace.
    return all(cell_value is None or str(cell_value).strip() == '' for cell_value in row_data)

# --- Core Data Reading and Table Identification Functions ---

def read_sheet_data(sheet):
    """
    Reads all data from an openpyxl sheet.
    Returns a list of lists, where each inner list represents a row.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet object from openpyxl.

    Returns:
        list: A list of lists, where each inner list contains cell_value.
    """
    data = []
    # iter_rows() is efficient for reading all cells
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)
    return data

def find_table_boundaries(sheet_data, table_keyword="Table"):
    """
    Finds the start and end row indices of tables within a sheet's data based on a keyword.
    A table block starts with a row containing the 'table_keyword'.
    The header is the next non-empty row after the keyword row.
    The table data continues until an empty row or the end of the sheet.

    Args:
        sheet_data (list): A list of lists, where each inner list contains cell values.
        table_keyword (str): The keyword to look for to identify the start of a table block.

    Returns:
        list: A list of (table_block_start_row, table_data_end_row, actual_header_row) tuples.
              All row indices are 0-based.
    """
    table_boundaries = []
    num_rows = len(sheet_data)
    r_idx = 0

    while r_idx < num_rows:
        row_content = " ".join([str(cell_value) for cell_value in sheet_data[r_idx] if cell_value is not None]).strip()

        # 1. Scan for "Table" keyword to find the start of a table block
        if table_keyword.lower() in row_content.lower():
            table_block_start_row = r_idx
            print(f"DEBUG: Found table keyword '{table_keyword}' at row {table_block_start_row}")

            # 2. Find Actual Header: Next non-empty row after the table_block_start_row
            actual_header_row = -1
            for h_idx in range(table_block_start_row + 1, num_rows):
                if not is_row_empty(sheet_data[h_idx]): # Use the helper from Part 1
                    actual_header_row = h_idx
                    print(f"DEBUG: Found actual header at row {actual_header_row}: {sheet_data[actual_header_row]}")
                    break
            
            if actual_header_row == -1:
                # No header found after table keyword, skip this block
                print(f"DEBUG: No actual header found after table keyword at row {table_block_start_row}. Skipping block.")
                r_idx += 1
                continue

            # 3. Find Table Data End: Stop at the next empty row OR when the next "Table" block starts
            table_data_end_row = actual_header_row
            for d_idx in range(actual_header_row + 1, num_rows):
                # If we encounter a new table keyword, end current table just before it
                next_row_text = " ".join([
                    str(cell_value) for cell_value in sheet_data[d_idx]
                    if cell_value is not None
                ]).strip().lower()
                if table_keyword.lower() in next_row_text:
                    table_data_end_row = d_idx - 1
                    print(f"DEBUG: Found start of next table at row {d_idx}; ending current table at {table_data_end_row}")
                    break

                # If we encounter an empty row, end the table just before it
                if is_row_empty(sheet_data[d_idx]):
                    table_data_end_row = d_idx - 1
                    print(f"DEBUG: Found end of data block at row {d_idx-1} (empty row at {d_idx})")
                    break

                # Otherwise, continue extending the table range
                table_data_end_row = d_idx
            
            print(f"DEBUG: Final table boundaries: Block start: {table_block_start_row}, Data end: {table_data_end_row}, Actual Header: {actual_header_row}")
            table_boundaries.append((table_block_start_row, table_data_end_row, actual_header_row))
            r_idx = table_data_end_row + 1 # Continue search after this table
        else:
            r_idx += 1 # Move to next row if no keyword found

    return table_boundaries

def extract_table_data(sheet_data, table_block_start_row, table_data_end_row, actual_header_row):
    """
    Extracts data for a specific table range into a pandas DataFrame.
    The 'table_block_start_row' (containing the "Table" keyword) is excluded from data.

    Args:
        sheet_data (list): A list of lists, where each inner list contains cell values.
        table_block_start_row (int): The 0-based row index where the "Table" keyword was found.
        table_data_end_row (int): The 0-based row index of the last data row in the table.
        actual_header_row (int): The 0-based row index of the actual header row.

    Returns:
        pandas.DataFrame: A DataFrame containing the extracted table data.
    """
    # Extract header row values from the actual header row
    header_values = [cell_value for cell_value in sheet_data[actual_header_row]]
    print(f"DEBUG: extract_table_data - Header values: {header_values}")
    # Filter out None or empty string headers to avoid unnamed columns, provide default if empty
    header_values = [h if h is not None and str(h).strip() != '' else f'Unnamed_{i}' for i, h in enumerate(header_values)]

    

    table_rows = []
    # Data rows start from the row AFTER the actual header row
    for r_idx in range(actual_header_row + 1, table_data_end_row + 1):
        row_values = [cell_value for cell_value in sheet_data[r_idx]]
        
        table_rows.append(row_values)
    
    # Create DataFrame
    if table_rows:
        # Use the length of the first row to determine the number of columns for data
        # This handles cases where header_values might be longer than actual data rows
        num_data_cols = len(table_rows[0])
        df = pd.DataFrame(table_rows, columns=header_values[:num_data_cols])
    else:
        # If no data rows, create an empty DataFrame with the expected headers
        df = pd.DataFrame(columns=header_values)

    return df

def extract_remarks(sheet_data, all_table_boundaries):
    """
    Extracts remarks/notes from a sheet, assuming they are outside identified table boundaries.
    Table boundaries now include the 'Table' keyword row, header row, and data rows.

    Args:
        sheet_data (list): A list of lists, where each inner list contains cell values.
        all_table_boundaries (list): A list of (table_block_start_row, table_data_end_row, actual_header_row) tuples for ALL tables in the sheet.

    Returns:
        list: A list of strings, where each string is a non-empty row identified as a remark.
    """
    remarks = []
    # Collect all row indices that are part of any identified table block (including the "Table" keyword row)
    all_table_block_rows = set()
    for table_block_start, table_data_end, actual_header in all_table_boundaries:
        # Include the "Table" keyword row and all rows up to the end of data
        for r_idx in range(table_block_start, table_data_end + 1):
            all_table_block_rows.add(r_idx)

    for r_idx, row in enumerate(sheet_data):
        if r_idx not in all_table_block_rows:
            row_text = " ".join([
                str(cell_value) for cell_value in row 
                if cell_value is not None and str(cell_value).strip() != ''
            ]).strip()
            if row_text:
                remarks.append(row_text)
    return remarks

def save_df_with_metadata(df, remarks, header_sheet_row, data_end_sheet_row, table_length, original_excel_path, table_name, output_dir):
    """
    Saves a DataFrame to CSV and appends remarks, line numbers, and table length to the end of the file.

    Args:
        df (pandas.DataFrame): The DataFrame to save.
        remarks (list): A list of remark strings to append.
        header_sheet_row (int): The 0-based row index of the table's header in the original sheet.
        data_end_sheet_row (int): The 0-based row index of the table's last data row in the original sheet.
        table_length (int): The number of data rows in the table (excluding header).
        original_excel_path (str): The full path to the original Excel file.
        table_name (str): The name of the table (e.g., 'Table1', 'Table3').
        output_dir (str): The directory where the CSV will be saved.
    """
    base_excel_name = os.path.splitext(os.path.basename(original_excel_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = os.path.join(output_dir, f"{base_excel_name}_{table_name}_{timestamp}.csv")

    df.to_csv(output_filename, index=False, encoding="utf-8")

    # Append metadata to the CSV file
    with open(output_filename, "a", encoding="utf-8") as f:
        f.write(f"\n--- Table Metadata for {table_name} ---\n")
        f.write(f"Header Row (0-indexed in sheet): {header_sheet_row}\n")
        f.write(f"Last Data Row (0-indexed in sheet): {data_end_sheet_row}\n")
        f.write(f"Number of Data Rows (Length of Table): {table_length}\n")

        if remarks:
            f.write("\n--- Remarks/Notes ---\n")
            for remark in remarks:
                f.write(remark + "\n")
    
    print(f"Successfully saved {table_name} to {output_filename}")

def process_excel_file(file_path):
    """
    Processes the specified Excel file according to the defined requirements.
    Orchestrates reading sheets, finding tables, extracting data, and saving.

    Args:
        file_path (str): The full path to the Excel file to process.
    """
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    print(f"Processing Excel file: {file_path}")
    # Load workbook, data_only=True ensures cell values are read, not formulas
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # --- Special handling for WO2021163344_extracted_250611.xlsx ---
    base_excel_name = os.path.splitext(os.path.basename(file_path))[0]
    if base_excel_name == "WO2021163344_extracted_250611":
        # Sheet 1 (index 0): drop first line, keep columns 1..3
        if len(workbook.worksheets) >= 1:
            s1_name = workbook.worksheets[0].title
            df1 = pd.read_excel(file_path, sheet_name=0, header=None)
            df1 = df1.iloc[1:, 0:3]
            save_df_with_metadata(df1, [], 0, len(df1), len(df1), file_path, s1_name, OUTPUT_DIR)

        # Sheet 2 (index 1): keep columns 1..4
        if len(workbook.worksheets) >= 2:
            s2_name = workbook.worksheets[1].title
            df2 = pd.read_excel(file_path, sheet_name=1, header=None)
            df2 = df2.iloc[:, 0:4]
            save_df_with_metadata(df2, [], 0, len(df2), len(df2), file_path, s2_name, OUTPUT_DIR)

        # Sheet 4 (index 3): drop first line, keep columns 1..5
        if len(workbook.worksheets) >= 4:
            s4_name = workbook.worksheets[3].title
            df4 = pd.read_excel(file_path, sheet_name=3, header=None)
            df4 = df4.iloc[1:, 0:5]
            save_df_with_metadata(df4, [], 0, len(df4), len(df4), file_path, s4_name, OUTPUT_DIR)

        # Done with special file
        return

    # --- Sheet 1 (index 0): Table 1 and Table 2 ---
    if len(workbook.worksheets) > 0:
        sheet1 = workbook.worksheets[0]
        sheet1_data = read_sheet_data(sheet1)
        sheet1_tables = find_table_boundaries(sheet1_data)
        print(f"DEBUG: Sheet 1 - Found tables: {sheet1_tables}")

        # Process up to two tables if present (do not require both to exist)
        if len(sheet1_tables) >= 1:
            table1_block_start, table1_data_end, table1_header = sheet1_tables[0]
            print(f"DEBUG: Processing Table1. Block start: {table1_block_start}, Data end: {table1_data_end}, Header: {table1_header}")
            table1_df = extract_table_data(sheet1_data, table1_block_start, table1_data_end, table1_header)
            table1_remarks = extract_remarks(sheet1_data, sheet1_tables)
            save_df_with_metadata(table1_df, table1_remarks, table1_header, table1_data_end, len(table1_df), file_path, 'Table1', OUTPUT_DIR)
        else:
            print(f"Warning: Could not find Table1 in the first sheet of {file_path}")

        if len(sheet1_tables) >= 2:
            table2_block_start, table2_data_end, table2_header = sheet1_tables[1]
            print(f"DEBUG: Processing Table2. Block start: {table2_block_start}, Data end: {table2_data_end}, Header: {table2_header}")
            table2_df = extract_table_data(sheet1_data, table2_block_start, table2_data_end, table2_header)
            table2_remarks = extract_remarks(sheet1_data, sheet1_tables)
            save_df_with_metadata(table2_df, table2_remarks, table2_header, table2_data_end, len(table2_df), file_path, 'Table2', OUTPUT_DIR)
        else:
            print(f"Warning: Could not find Table2 in the first sheet of {file_path}")
    else:
        print(f"Warning: No sheets found in {file_path}")

    # --- Sheet 2 (index 1): Table 3 ---
    if len(workbook.worksheets) > 1:
        sheet2 = workbook.worksheets[1]
        sheet2_data = read_sheet_data(sheet2)
        sheet2_tables = find_table_boundaries(sheet2_data)
        print(f"DEBUG: Sheet 2 - Found tables: {sheet2_tables}")

        if len(sheet2_tables) >= 1:
            table3_block_start, table3_data_end, table3_header = sheet2_tables[0]
            print(f"DEBUG: Processing Table3. Block start: {table3_block_start}, Data end: {table3_data_end}, Header: {table3_header}")
            table3_df = extract_table_data(sheet2_data, table3_block_start, table3_data_end, table3_header)
            table3_remarks = extract_remarks(sheet2_data, sheet2_tables) # Pass all tables for remarks
            save_df_with_metadata(table3_df, table3_remarks, table3_header, table3_data_end, len(table3_df), file_path, 'Table3', OUTPUT_DIR)
        else:
            print(f"Warning: Could not find Table3 in the second sheet of {file_path}")

    # --- Sheet 3 (index 2): Table 4 ---
    if len(workbook.worksheets) > 2:
        sheet3 = workbook.worksheets[2]
        sheet3_data = read_sheet_data(sheet3)
        sheet3_tables = find_table_boundaries(sheet3_data)
        print(f"DEBUG: Sheet 3 - Found tables: {sheet3_tables}")

        if len(sheet3_tables) >= 1: # Assuming Table 4 is the first table found
            table4_block_start, table4_data_end, table4_header = sheet3_tables[0]
            print(f"DEBUG: Processing Table4. Block start: {table4_block_start}, Data end: {table4_data_end}, Header: {table4_header}")
            table4_df = extract_table_data(sheet3_data, table4_block_start, table4_data_end, table4_header)
            table4_remarks = extract_remarks(sheet3_data, sheet3_tables) # Pass all tables for remarks
            save_df_with_metadata(table4_df, table4_remarks, table4_header, table4_data_end, len(table4_df), file_path, 'Table4', OUTPUT_DIR)
        else:
            print(f"Warning: Could not find Table4 in the third sheet of {file_path}")

    # --- Sheet 4 (index 3): Table 5 through 14 ---
    if len(workbook.worksheets) > 3:
        sheet4 = workbook.worksheets[3]
        sheet4_data = read_sheet_data(sheet4)
        sheet4_tables = find_table_boundaries(sheet4_data)
        print(f"DEBUG: Sheet 4 - Found tables: {sheet4_tables}")

        for i, (table_block_start, table_data_end, actual_header) in enumerate(sheet4_tables):
            if i >= 0 and i < 10: # Assuming tables 5-14 are the first 10 tables found
                table_name = f'Table{i+5}'
                print(f"DEBUG: Processing {table_name}. Block start: {table_block_start}, Data end: {table_data_end}, Header: {actual_header}")
                table_df = extract_table_data(sheet4_data, table_block_start, table_data_end, actual_header)
                table_remarks = extract_remarks(sheet4_data, sheet4_tables) # Pass all tables for remarks
                save_df_with_metadata(table_df, table_remarks, actual_header, table_data_end, len(table_df), file_path, table_name, OUTPUT_DIR)
            else:
                print(f"Warning: More tables found in the fourth sheet than expected (Tables 5-14). Skipping table {i+1}.")

    # --- Sheet 5 (index 4): Last sheet, no specific table ---
    if len(workbook.worksheets) > 4: # Check if 5th sheet exists
        last_sheet = workbook.worksheets[4] # Access 5th sheet by index
        last_sheet_name = last_sheet.title # Get title for filename
        last_sheet_data = read_sheet_data(last_sheet)
        
        # Convert all data from the last sheet into a DataFrame
        # Assuming the first row is header for this general conversion
        if last_sheet_data:
            header_row_data = [cell_value for cell_value in last_sheet_data[0]]
            data_rows = [[cell_value for cell_value in r] for r in last_sheet_data[1:]]
            
            # Ensure columns match data length, handle potential empty data_rows
            if data_rows and header_row_data and len(data_rows[0]) == len(header_row_data):
                last_sheet_df = pd.DataFrame(data_rows, columns=header_row_data)
            elif not data_rows and header_row_data: # Empty data but has header
                last_sheet_df = pd.DataFrame(columns=header_row_data)
            else: # Fallback for no header or mismatch
                last_sheet_df = pd.DataFrame(data_rows)

            print(f"DEBUG: Processing last sheet ({last_sheet_name}). Total rows: {len(last_sheet_data)}")
            save_df_with_metadata(last_sheet_df, [], 0, len(last_sheet_data) - 1, len(last_sheet_df), file_path, last_sheet_name, OUTPUT_DIR) # No specific remarks to extract, save all
        else:
            print(f"Warning: Sheet '{last_sheet_name}' (index 4) is empty.")

if __name__ == "__main__":
    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_DIR, exist_ok=True) # Use exist_ok=True to prevent error if dir exists

    # This script is designed for a specific file
    current_dir = Path(__file__).resolve().parent
    excel_file_path = current_dir / EXCEL_FILE_NAME

    if os.path.exists(excel_file_path):
        process_excel_file(str(excel_file_path))
    else:
        print(f"Error: The specified Excel file '{EXCEL_FILE_NAME}' was not found in the current directory.")
        print("Please ensure the file is in the same directory as the script, or update EXCEL_FILE_NAME.")