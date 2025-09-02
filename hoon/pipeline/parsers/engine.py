import hashlib
import os
import pandas as pd
from openpyxl import load_workbook


def read_excel(path, sheet, header=None, dtype_str=True):
    """Args: path(str), sheet(str|int), header(int|None), dtype_str(bool) -> DataFrame

    Excel 시트를 읽어 DataFrame을 반환한다. dtype_str=True면 문자열로 강제.
    """
    kwargs = {}
    if dtype_str:
        kwargs["dtype"] = str
    return pd.read_excel(path, sheet_name=sheet, header=header, engine="openpyxl", **kwargs)


def extract_block(df, a1_range):
    """Args: df(DataFrame), a1_range(str) -> DataFrame

    A1 표기("A2:D10")로 지정된 블록을 잘라 반환한다.
    """
    import re
    # 간단한 A1 파서 (열:문자→인덱스, 행:숫자)
    def col_to_idx(col):
        col = col.upper()
        acc = 0
        for c in col:
            acc = acc * 26 + (ord(c) - ord("A") + 1)
        return acc - 1

    m = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", str(a1_range).strip())
    if not m:
        raise ValueError("잘못된 A1 범위: %s" % a1_range)
    c1, r1, c2, r2 = m.groups()
    r1, r2 = int(r1) - 1, int(r2) - 1
    x1, x2 = col_to_idx(c1), col_to_idx(c2)
    return df.iloc[r1 : r2 + 1, x1 : x2 + 1].copy()


def apply_header(block, header_row_offset, nrows=1):
    """Args: block(DataFrame), header_row_offset(int), nrows(int) -> DataFrame

    블록 상단에서 헤더 행을 추출해 컬럼명을 적용한다.
    """
    hdr = block.iloc[header_row_offset : header_row_offset + nrows].astype(str)
    data = block.iloc[header_row_offset + nrows :].reset_index(drop=True)
    if nrows == 1:
        data.columns = hdr.iloc[0].tolist()
    else:
        data.columns = pd.MultiIndex.from_frame(hdr.reset_index(drop=True).T)
    return data


def melt_table(df, id_cols, value_cols):
    """Args: df(DataFrame), id_cols(list), value_cols(list) -> DataFrame

    와이드 테이블을 롱 포맷으로 변환한다.
    """
    return df.melt(id_vars=id_cols, value_vars=value_cols, var_name="variable", value_name="value")


def matrix_to_long(block, cfg):
    """Args: block(DataFrame), cfg(dict) -> DataFrame

    2단 헤더(패널/세포주)를 롱 포맷으로 변환한다.
    cfg 필수 키: panel_row_offset, cellline_row_offset, data_row_start, id_col, panels
    선택 키: value_parser:{unit_default, rules(list of "regex -> assignments")}
    """
    panel_row = int(cfg.get("panel_row_offset", 0))
    cell_row = int(cfg.get("cellline_row_offset", panel_row + 1))
    data_start = int(cfg.get("data_row_start", cell_row + 1))
    id_col = int(cfg.get("id_col", 0))
    panels = cfg.get("panels", []) or []

    # 헤더 행 추출
    header_panel = block.iloc[panel_row]
    header_cell = block.iloc[cell_row]
    rows = []

    def parse_by_rules(text, vp):
        from .measures import parse_measure
        if text is None:
            return None, None, None
        s = str(text).strip()
        if s == "":
            return None, None, None
        # 정규식 매칭 규칙 우선
        rules = ((vp or {}).get("rules") or [])
        for rule in rules:
            if "->" not in rule:
                continue
            pat, rhs = rule.split("->", 1)
            import re
            m = re.search(pat.strip(), s)
            if not m:
                continue
            # 기본 파싱으로 채우고, rhs 지정이 있으면 오버라이드
            q, v, u = parse_measure(s, default_unit=(vp or {}).get("unit_default"))
            # 간단한 키: 'qualifier:\'>'', 'value:\1', 'unit:uM'
            kvs = {}
            for part in rhs.split(','):
                part = part.strip()
                if not part:
                    continue
                if ":" not in part:
                    continue
                k, vexpr = part.split(":", 1)
                k = k.strip()
                vexpr = vexpr.strip()
                if vexpr.startswith("'") and vexpr.endswith("'"):
                    kvs[k] = vexpr[1:-1]
                elif vexpr.startswith("\"") and vexpr.endswith("\""):
                    kvs[k] = vexpr[1:-1]
                elif vexpr.startswith("\\") and len(vexpr) > 1 and vexpr[1:].isdigit():
                    # 역참조 숫자 -> 그룹
                    idx = int(vexpr[1:])
                    kvs[k] = m.group(idx)
                elif vexpr.startswith("\\") and len(vexpr) > 1:
                    # 그대로 텍스트
                    kvs[k] = vexpr[1:]
                else:
                    kvs[k] = vexpr
            q = kvs.get("qualifier", q)
            v = kvs.get("value", v)
            u = kvs.get("unit", u)
            return q, v, u
        # 규칙이 없거나 매칭 실패 시 일반 파서
        return parse_measure(s, default_unit=(vp or {}).get("unit_default"))

    vp = cfg.get("value_parser") or {}
    for r in range(data_start, len(block)):
        rid = block.iloc[r, id_col]
        for p in panels:
            pname = str(p.get("name", "")).strip()
            for c in p.get("cols", []) or []:
                c = int(c)
                cell_line = str(header_cell.iloc[c]) if c < len(header_cell) else ""
                raw = block.iloc[r, c] if c < block.shape[1] else None
                q, v, u = parse_by_rules(raw, vp)
                rows.append({
                    "row_id": rid,
                    "panel": pname,
                    "cell_line": cell_line,
                    "value_raw": raw,
                    "qualifier": q,
                    "value": v,
                    "unit": u,
                })
    import pandas as pd
    return pd.DataFrame(rows, columns=["row_id","panel","cell_line","value_raw","qualifier","value","unit"]) if rows else pd.DataFrame(columns=["row_id","panel","cell_line","value_raw","qualifier","value","unit"]) 


def read_cell_colors(path, sheet):
    """Args: path(str), sheet(str|int) -> dict

    셀 색상을 {(row_index0,col_index0): rgb_hex} 형태로 반환한다.
    색상 정보가 없으면 비어있는 dict 반환.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[int(sheet)]
    colors = {}
    for r_idx, row in enumerate(ws.iter_rows()):
        for c_idx, cell in enumerate(row):
            fill = getattr(cell, 'fill', None)
            if not fill or not getattr(fill, 'fgColor', None):
                continue
            fg = fill.fgColor
            rgb = getattr(fg, 'rgb', None)
            if rgb and len(rgb) == 8:
                # ARGB -> RGB
                colors[(r_idx, c_idx)] = rgb[-6:].upper()
            elif getattr(fg, 'type', None) == 'rgb' and hasattr(fg, 'value'):
                val = getattr(fg, 'value')
                if isinstance(val, str) and len(val) >= 6:
                    colors[(r_idx, c_idx)] = val[-6:].upper()
    return colors


def sanitize_strings(df):
    """Args: df(DataFrame) -> DataFrame

    오브젝트/문자열 컬럼에서 널바이트(\x00) 제거 및 양끝 공백 제거.
    원본을 변경하지 않고 정제된 복사본을 반환.
    """
    out = df.copy()
    # 중복 컬럼명이 있을 수 있어 위치 기반으로 처리
    for i in range(out.shape[1]):
        col = out.iloc[:, i]
        # object처럼 취급 가능한 경우
        try:
            out.iloc[:, i] = col.map(lambda x: None if x is None else str(x).replace('\x00','').strip())
        except Exception:
            # 숫자 등 매핑 불가 타입은 그대로 둠
            pass
    return out


def file_sha256(path):
    """Args: path(str) -> str

    파일의 SHA256 해시를 16진 문자열로 반환.
    """
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()
